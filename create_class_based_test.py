from openpyxl import Workbook

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "IP Addresses"

# Add header
ws['A1'] = 'ip'

# Add sample IP addresses demonstrating class-based grouping
ips = [
    # Class A (1-127): Should group by first octet
    # Group: 10.x.x.x (5 IPs -> 1 query)
    '10.0.0.1',
    '10.0.1.1',
    '10.1.0.1',
    '10.5.10.20',
    '10.255.255.255',
    
    # Class B (128-191): Should group by first two octets
    # Group: 172.16.x.x (3 IPs -> 1 query)
    '172.16.0.1',
    '172.16.1.1',
    '172.16.255.255',
    
    # Group: 172.17.x.x (2 IPs -> 1 query)
    '172.17.0.1',
    '172.17.1.1',
    
    # Class C (192-223): Should group by first three octets
    # Group: 192.168.1.x (4 IPs -> 1 query)
    '192.168.1.1',
    '192.168.1.100',
    '192.168.1.200',
    '192.168.1.254',
    
    # Group: 192.168.2.x (2 IPs -> 1 query)
    '192.168.2.1',
    '192.168.2.100',
    
    # Single IPs (no grouping)
    '8.8.8.8',      # Google DNS
    '1.1.1.1',      # Cloudflare DNS
]

for idx, ip in enumerate(ips, start=2):
    ws[f'A{idx}'] = ip

# Save the file
wb.save('test_class_based_grouping.xlsx')

print("Test Excel file created: test_class_based_grouping.xlsx")
print(f"\nTotal IPs: {len(ips)}")
print(f"Expected API calls: 7")
print(f"  - Class A (10.x.x.x): 5 IPs -> 1 query")
print(f"  - Class B (172.16.x.x): 3 IPs -> 1 query")
print(f"  - Class B (172.17.x.x): 2 IPs -> 1 query")
print(f"  - Class C (192.168.1.x): 4 IPs -> 1 query")
print(f"  - Class C (192.168.2.x): 2 IPs -> 1 query")
print(f"  - Single IPs: 2 -> 2 queries")
print(f"\nOptimization: {len(ips) - 7} API calls saved! ({((len(ips) - 7) / len(ips) * 100):.1f}%)")

print("\n\nDetailed IP Groups:")
print("\n  Class A - 10.x.x.x (5 IPs):")
for ip in ips[:5]:
    print(f"    - {ip}")

print("\n  Class B - 172.16.x.x (3 IPs):")
for ip in ips[5:8]:
    print(f"    - {ip}")

print("\n  Class B - 172.17.x.x (2 IPs):")
for ip in ips[8:10]:
    print(f"    - {ip}")

print("\n  Class C - 192.168.1.x (4 IPs):")
for ip in ips[10:14]:
    print(f"    - {ip}")

print("\n  Class C - 192.168.2.x (2 IPs):")
for ip in ips[14:16]:
    print(f"    - {ip}")

print("\n  Single IPs (2):")
for ip in ips[16:]:
    print(f"    - {ip}")
