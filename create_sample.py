from openpyxl import Workbook

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "IP Addresses"

# Add header
ws['A1'] = 'ip'

# Add sample IP addresses
ips = [
    '8.8.8.8',
    '1.1.1.1',
    '208.67.222.222',
    '9.9.9.9',
    '64.6.64.6'
]

for idx, ip in enumerate(ips, start=2):
    ws[f'A{idx}'] = ip

# Save the file
wb.save('sample_ips.xlsx')

print("Sample Excel file created: sample_ips.xlsx")
print(f"Total IPs: {len(ips)}")
for ip in ips:
    print(f"  - {ip}")
