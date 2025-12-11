from openpyxl import Workbook

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "IP Addresses"

# Add header
ws['A1'] = 'ip'

# Add sample IP addresses with grouping
# Group 1: 192.168.x.x (5 IPs - should query only 1)
# Group 2: 10.0.x.x (3 IPs - should query only 1)
# Single IPs: 8.8.8.8, 1.1.1.1 (should query both)
ips = [
    # Group 1: 192.168.x.x
    '192.168.1.1',
    '192.168.1.100',
    '192.168.1.200',
    '192.168.2.1',
    '192.168.3.50',
    
    # Group 2: 10.0.x.x
    '10.0.0.1',
    '10.0.1.1',
    '10.0.2.1',
    
    # Single IPs
    '8.8.8.8',
    '1.1.1.1',
]

for idx, ip in enumerate(ips, start=2):
    ws[f'A{idx}'] = ip

# Save the file
wb.save('test_grouped_ips.xlsx')

print("Test Excel file created: test_grouped_ips.xlsx")
print(f"\nTotal IPs: {len(ips)}")
print(f"Expected API calls: 4 (1 for 192.168 group, 1 for 10.0 group, 2 single IPs)")
print(f"Optimization: {len(ips) - 4} API calls saved!\n")

print("IP Groups:")
print("  192.168.x.x: 5 IPs")
for ip in ips[:5]:
    print(f"    - {ip}")
print("\n  10.0.x.x: 3 IPs")
for ip in ips[5:8]:
    print(f"    - {ip}")
print("\n  Single IPs: 2")
for ip in ips[8:]:
    print(f"    - {ip}")
